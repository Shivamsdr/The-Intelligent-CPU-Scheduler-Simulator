import random
from openpyxl import load_workbook
import pandas as pd
from openpyxl import Workbook
from algorithms import fcfs, sjf, round_robin as rr, priority_scheduling, srtf
#from src.suggest_algorithm import suggest_algorithm
from suggest2 import suggest_algorithm

'''SYNTHETIC POPULATIOn'''
def generate_processes(num_processes):
    processes = []
    for i in range(num_processes):
        arrival_time = random.randint(0, 15)  
        burst_time = random.randint(1, 15)   
        priority= 0 #random.randint(0, 5) if random.random() > 0.5 else 0  
        processes.append({
            'id': str(i),
            'arrival_time': arrival_time,
            'burst_time': burst_time,
            'priority': priority,
            'initial_burst_time': burst_time
        })
    return processes

def export_to_excel(processes, filename=r'C:\Users\oms1n\OneDrive\Desktop\CS\SamplePr.xlsx'):
    df = pd.DataFrame(processes)
    df.to_excel(filename, index=False)
    print(f"Processes saved to {filename}")
    
''' RANDOM SAMPLE '''
def gen_sample(row,maxsize=50):
    df = pd.read_excel(r'C:\Users\oms1n\OneDrive\Desktop\CS\SamplePr.xlsx',usecols=['id', 'arrival_time',
                                                                                    'burst_time', 'priority'])
    processes = df.to_dict('records')
    for i in range(row):
        # creating random sample intances
        random_sample = random.sample(processes, random.randint(1, min(maxsize, len(processes))))

        #run all algos
        results = {
        'fcfs': calculate_metrics(fcfs(random_sample)),
        'sjf': calculate_metrics(sjf(random_sample)),
        'srtf': calculate_metrics(srtf(random_sample)),
        'rr': calculate_metrics(rr(random_sample, quantum=4)),
        'priority': calculate_metrics(priority_scheduling(random_sample))
        }

        # Making Columns
        sugg_algo = suggest_algorithm(random_sample)
        best_algo_tt = min(results, key=lambda algo: results[algo]['avg_turnaround_time'])
        best_algo_wt = min(results, key=lambda algo: results[algo]['avg_waiting_time'])
        n = len(random_sample)
        p = ', '.join([f"P-{proc['id']}({proc['arrival_time']},{proc['burst_time']})" for proc in random_sample])
        
        append_to_excel(p, n, sugg_algo, best_algo_tt,best_algo_wt, results)
        # d = 118 // row
        # pp = min(d, 64)  # Cap pp at 64
        # for _ in range(pp):
        #     print(".", end="", flush=True)

        progress = int(((i + 1) / row) * 100)  # Calculate percentage
        bar = "#" * (progress // 2) + "-" * (50 - (progress // 2))  # Create the progress bar (50 characters wide)
        print(f"\r[{bar}] {progress}%", end="", flush=True)  # Print the progress bar dynamically

    print(f" ✅\n{row} Random sample[s] saved")

def append_to_excel(p,n,sugg_algo,best_algo_tt,best_algo_wt,results):
    output_file = r'C:\Users\oms1n\OneDrive\Desktop\CS\RandSamples.xlsx'
    try:
        workbook = load_workbook(output_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        # Add headers if creating a new file
        sheet.append(['p', 'n', 'sugg_algo','best_algo_tt', 'fcfs', 'sjf', 'srtf', 'rr', 'priority',
                      '','best_algo_wt', 'fcfs', 'sjf', 'srtf', 'rr', 'priority'])

    # Append data to the sheet
    sheet.append([p, n, sugg_algo,best_algo_tt,
                  results['fcfs']['avg_turnaround_time'],
                  results['sjf']['avg_turnaround_time'],
                  results['srtf']['avg_turnaround_time'],
                  results['rr']['avg_turnaround_time'],
                  results['priority']['avg_turnaround_time'],
                  '',
                  best_algo_wt,
                  results['fcfs']['avg_waiting_time'],
                  results['sjf']['avg_waiting_time'],
                  results['srtf']['avg_waiting_time'],
                  results['rr']['avg_waiting_time'],
                  results['priority']['avg_waiting_time']])
    # Save 
    workbook.save(output_file)

#seperately made here for best algo selection
def calculate_metrics(result): 
    #Calculate metrics for result.
    total_turnaround_time = sum(p['turnaround_time'] for p in result)
    total_waiting_time = sum(p['waiting_time'] for p in result)
    n = len(result)
    return {
        'avg_turnaround_time': total_turnaround_time,
        'avg_waiting_time': total_waiting_time
    }

# def BEST_stat():
#     df = pd.read_excel(r'C:\Users\oms1n\OneDrive\Desktop\CS\RandSamples.xlsx')
#     print("Best Algorithm Statistics:")
#     print(df['best_algo'].value_counts())

if __name__ == '__main__':
    pop = input("Generate new population? 1/0 ")
    if(pop=='y' or pop == 1):
        processes = generate_processes(1000)
        export_to_excel(processes) #overwrites population
    x = int(input("Enter number of random samples to generate: "))
    gen_sample(x) #append new sample